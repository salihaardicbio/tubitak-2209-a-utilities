"""
Convert Foldseek search-server JSON exports into single-sheet Excel workbooks.

All hits from every searched database (BFVD, afdb50, pdb100, etc.) are
collected into one sheet, with a "database" column identifying the source,
sorted by database then e-value.

CAUTION: This script processes every .json file found in SEARCH_FOLDER
(including subfolders). It assumes each one is a Foldseek server export
with the specific structure (a list containing a dict with a "results"
key). If the folder contains unrelated .json files, this will either
throw an error on that file (and skip it, printing a message) or, in
rare cases where a file coincidentally shares similar keys, produce a
malformed/empty sheet. I have not tested this against non-Foldseek JSON,
so review the printed summary at the end and spot-check the output
before relying on it, especially in a folder with mixed file types.

Batch mode: edit SEARCH_FOLDER at the bottom of this script, then run it.
It writes a matching .xlsx (same base name) for every .json file found,
to OUTPUT_FOLDER. The output filename (before ".xlsx") is also used as
the "query" value for every row.

    python foldseek_json_to_xlsx.py

--include-coords : also export the raw 3D CA coordinate string (tCa) for
                    the target structure. Off by default since it's a very
                    long string and not usable directly in a spreadsheet.
"""
import json
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_COLS = [
    'database', 'query', 'target', 'seqId', 'alnLength', 'missmatches', 'gapsopened',
    'qStartPos', 'qEndPos', 'dbStartPos', 'dbEndPos', 'prob', 'eval', 'score',
    'qLen', 'dbLen', 'taxId', 'taxName', 'description', 'qAln', 'dbAln', 'tSeq',
]

COL_WIDTHS = {
    'database': 14, 'query': 10, 'target': 16, 'seqId': 8, 'alnLength': 10,
    'missmatches': 11, 'gapsopened': 11, 'qStartPos': 10, 'qEndPos': 9,
    'dbStartPos': 11, 'dbEndPos': 9, 'prob': 7, 'eval': 10, 'score': 8,
    'qLen': 7, 'dbLen': 7, 'taxId': 9, 'taxName': 22, 'description': 22,
    'qAln': 40, 'dbAln': 40, 'tSeq': 40, 'tCa': 20,
}


def convert(json_path: str, xlsx_path: str, include_coords: bool = False):
    with open(json_path) as f:
        data = json.load(f)

    item = data[0]
    cols = BASE_COLS + (['tCa'] if include_coords else [])

    job_name = os.path.splitext(os.path.basename(xlsx_path))[0]

    rows = []
    for r in item['results']:
        db = r['db']
        for aln_list in r['alignments'].values():
            for aln in aln_list:
                row = {c: aln.get(c, '') for c in cols if c not in ('database', 'query')}
                row['database'] = db
                row['query'] = job_name
                rows.append(row)

    df = pd.DataFrame(rows, columns=cols)
    if not df.empty:
        df['eval'] = df['eval'].astype(float)
        df = df.sort_values(['database', 'eval'])

    wb = Workbook()
    ws = wb.active
    ws.title = 'Foldseek hits'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='1F4E78')

    ws.append(cols)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for _, row in df.iterrows():
        ws.append(list(row))

    ws.freeze_panes = 'A2'
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(c, 12)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name='Arial', size=10)

    wb.save(xlsx_path)
    print(f'Saved {xlsx_path}')


if __name__ == '__main__':
    # --- Edit this -----------------------------------------------------
    SEARCH_FOLDER = 'C:/Users/salih/Desktop/FoldSeek Veri Tabanı'   # folder to search (includes subfolders)
    INCLUDE_COORDS = False
    OUTPUT_FOLDER = 'C:/Users/salih/Desktop/FoldSeek Veri Tabanı/outputs'
    # ---------------------------------------------------------------------

    json_paths = []
    for root, _dirs, files in os.walk(SEARCH_FOLDER):
        for f in files:
            if f.lower().endswith('.json'):
                json_paths.append(os.path.join(root, f))

    if not json_paths:
        print(f'No .json files found in {SEARCH_FOLDER}')

    converted, failed = [], []
    for json_path in json_paths:
        filename = os.path.basename(json_path)
        out_name = os.path.splitext(filename)[0] + '.xlsx'
        out_path = os.path.join(OUTPUT_FOLDER, out_name)
        try:
            convert(json_path, out_path, INCLUDE_COORDS)
            converted.append(filename)
        except Exception as e:
            print(f'FAILED: {filename} ({e})')
            failed.append(filename)

    print(f'\nDone. Converted: {len(converted)}, Failed: {len(failed)}')
    if failed:
        print('Failed files (likely not Foldseek JSON exports):', ', '.join(failed))
